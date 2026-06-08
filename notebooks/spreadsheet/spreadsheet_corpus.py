"""
A corpus example for tabular data in a spreadsheet.

This is currently a minimum viable example, and is specific to a particular kind of
spreadsheet.

Limitations:

This isn't appropriate for large datasets:

1. Each process will hold a complete copy of the spreadsheet and will create memory
pressure if you have a large spreadsheet
2. You can only fit ~1million rows in a spreadsheet
3. Access to rows in the spreadsheet can be slower than data formats that are more
   appropriate for high performance analytics.

"""

import asyncio
import collections
import dataclasses as dc
import os
import pathlib
import random
import re
import typing

from openpyxl import load_workbook
from loky import get_reusable_executor
from tinyhtml import h

from hyperreal.corpus import HyperrealCorpus
from hyperreal.field_types import ValueSequence, Value, RangeEncodableValue
from hyperreal.index_core import HyperrealIndex, TableFilter
from hyperreal.web_server import serve_index


def extract_header(sheet) -> list[str]:
    """
    Extract the contiguous header columns from the first row of the sheet.

    """
    first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    headers = []

    if first_row:

        for val in first_row[0]:

            if isinstance(val, str):
                headers.append(val)
            else:
                break

    return headers


def identify_spreadsheet_tables(
    spreadsheet_path: pathlib.Path | str,
) -> dict[str, list[str]]:
    """
    Identify tables, columns and candidate joins in a give spreadsheet.

    """

    sheet_columns = dict()

    wb = load_workbook(spreadsheet_path)

    for sheet in wb:

        sheet_name = sheet.title

        headers = extract_header(sheet)

        if headers:
            sheet_columns[sheet_name] = extract_header(sheet)

    join_check = [(name, set(columns)) for name, columns in sheet_columns.items()]

    join_candidates = collections.defaultdict(dict)

    for i, (i_name, i_columns) in enumerate(join_check):

        for j_name, j_columns in join_check[i + 1 :]:

            inter_columns = i_columns & j_columns

            if inter_columns:
                join_candidates[i_name][j_name] = inter_columns
                join_candidates[j_name][i_name] = inter_columns

    return sheet_columns, join_candidates


boundary_regex = re.compile(r"\b|\s+")
display_regex = re.compile(r"(\b|\s+)")


def tokenise(text):
    return [t for t in boundary_regex.split(text.lower()) if t]


def display_tokenise(text):
    boundaries = [t for t in display_regex.split(text) if t]

    token_start = 0

    tokens = []

    for token_idx, part in enumerate(boundaries):
        # If it doesn't match our token boundary, it's a token and needs to be appended,
        # along with any previous split boundary tokens.
        if not display_regex.fullmatch(part):
            tokens.append("".join(boundaries[token_start : token_idx + 1]))
            token_start = token_idx + 1

    return tokens


@dc.dataclass
class XslxTranscriptCorpus(HyperrealCorpus):

    spreadsheet: pathlib.Path | str | bytes
    text_sheet: str = "turn"
    text_columns: list[str] = dc.field(default_factory=list)
    tokeniser: typing.Callable[[str], list[str]] = tokenise
    display_tokeniser: typing.Callable[[str], list[str]] = display_tokenise

    text_docs: list[tuple] = dc.field(init=False)
    text_col_idx: list[int] = dc.field(init=False)
    header_idx: list[tuple] = dc.field(init=False)
    lookup_tables: dict[dict] = dc.field(init=False)

    def __post_init__(self):

        wb = load_workbook(self.spreadsheet)

        sheet = wb[self.text_sheet]

        header = extract_header(sheet)
        self.header_idx = {col: i for i, col in enumerate(header)}

        all_rows = sheet.values
        # Skip the header
        next(all_rows)

        for col in self.text_columns:
            if col not in self.header_idx:
                raise ValueError(
                    f"Text column {col} not found in sheet {self.text_sheet}"
                )

        # We'll store the docs as a simple in memory list for now. There's only so many
        # documents we can fit in a spreadsheet after all.
        self.text_docs = []

        for row in all_rows:
            self.text_docs.append(row)

        # Next step: pull out the other tables as little in memory lookups based on the
        # selected keys. This is currently hard coded rather configurable while I work
        # out the interface.

        lookup_table_config = {
            "speaker_role": ("speaker_code",),
            "transcript_file": ("source_file",),
        }

        self.lookup_tables = {sheetname: dict() for sheetname in lookup_table_config}

        for sheetname, key_columns in lookup_table_config.items():

            lookup_table = dict()

            sheet = wb[sheetname]
            all_rows = sheet.values
            header = next(all_rows)
            key_col_idx = [header.index(col) for col in key_columns]

            for row in all_rows:
                lookup_table[tuple(row[col_idx] for col_idx in key_col_idx)] = row

            self.lookup_tables[sheetname] = lookup_table

    def __len__(self):
        return len(self.text_docs)

    def all_doc_keys(self):
        return range(len(self.text_docs))

    def docs(self, doc_keys):
        for key in doc_keys:

            doc_turn = self.text_docs[key]
            role = self.lookup_tables["speaker_role"][(doc_turn[3],)][1]
            interview_type, location = self.lookup_tables["transcript_file"][
                (doc_turn[0],)
            ][-2:]

            yield key, tuple((*doc_turn, role, interview_type, location))

    def doc_to_features(self, doc):
        doc_features = {
            col: ValueSequence(self.tokeniser(doc[self.header_idx[col]] or ""))
            for col in self.text_columns
        }

        doc_features["role"] = Value(doc[-3])
        doc_features["interview_type"] = Value(doc[-2])
        doc_features["location"] = Value(doc[-1])

        return doc_features

    def render_turn(self, doc, highlight_features=None):

        if highlight_features:
            match_features = {
                value for field, value in highlight_features if field == "transcription"
            }

            tokens = self.tokeniser(doc[4])
            display_tokens = self.display_tokeniser(doc[4])

            display_text = h("span")(
                (
                    display_token
                    if token not in match_features
                    else h("mark")(display_token)
                )
                for token, display_token in zip(tokens, display_tokens)
            )

        else:
            display_text = h("span")(doc[4])

        return h("div", klass="cluster")(h("em")(doc[3]), display_text)

    def html_search_results(self, doc_keys, highlight_features=None):

        search_hits = []

        for key in doc_keys:

            doc = list(self.docs([key]))[0][1]

            header_vals = [doc[-2], doc[-1], doc[-3]]
            header = h("div", klass="cluster result-header")(
                h("em")(val) for val in header_vals
            )

            context_range = range(max(0, key - 1), min(key + 2, len(self.text_docs)))

            turns = [
                self.render_turn(doc, highlight_features=highlight_features)
                for _, doc in self.docs(context_range)
            ]

            search_hits.append(h("li", klass="search-hit stack")(header, *turns))

        return h("ul", klass="stack search-results")(search_hits)

    def features_to_html_concordance(
        self, doc_features, display_features, highlight_features
    ):
        return None

    extra_css = """
        .result-header {
            font-weight: bold;
        }
    """


def serve_spreadsheet(spreadsheet):

    pool = get_reusable_executor()
    spreadsheet_corp = XslxTranscriptCorpus(
        spreadsheet=spreadsheet, text_sheet="turn", text_columns=["transcription"]
    )
    spreadsheet_idx = HyperrealIndex("spreadsheet_index.db", spreadsheet_corp, pool)

    print("Indexing documents")
    spreadsheet_idx.rebuild()

    spreadsheet_idx.search_fields = {"transcription": tokenise}

    # Configure the interface facets and search
    spreadsheet_idx.facets = [
        (
            "Speaker Role",
            spreadsheet_idx.field_features("role", min_docs=1),
            TableFilter(order_by="hits", first_k=20, keep_above=0),
        ),
        (
            "Interview Type",
            spreadsheet_idx.field_features("interview_type", min_docs=1),
            TableFilter(order_by="hits", keep_above=0),
        ),
        (
            "Location",
            spreadsheet_idx.field_features("location", min_docs=1),
            TableFilter(order_by="hits", keep_above=0),
        ),
    ]

    print("Creating feature clusters")

    clustering = spreadsheet_idx.plugins["feature_clusters"]

    # Set the state of the RNG to a consistent point.
    spreadsheet_idx.random_state = random.Random(42)

    # Initialise with a random clustering
    random_clustering = clustering.initialise_random_clustering(
        64, min_docs=3, include_fields=["transcription"]
    )

    clustering.replace_clusters(random_clustering)

    # Refine the clustering for a small number of iterations - we could go for
    # longer, but it usually doesn't matter as you'll spend the same amount of time
    # examining the output either way.
    clustering.refine_clustering(iterations=50)

    jupyter_hub_service_prefix = os.environ.get("JUPYTERHUB_SERVICE_PREFIX", "/")
    url_base = jupyter_hub_service_prefix + "proxy/absolute/9999"

    loop = asyncio.get_running_loop()
    task = loop.create_task(serve_index(spreadsheet_idx, base_path=url_base))

    display_link = h("a", href=url_base + "/browse/")("Browse your spreadsheet")
    display(display_link)
