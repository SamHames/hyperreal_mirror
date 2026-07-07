"""
User interface for configuring and launching a view of a tabular corpus.

"""

import pathlib
from io import BytesIO

import polars
from ipywidgets import (
    FileUpload,
    Layout,
    Output,
    Button,
    VBox,
    SelectMultiple,
    Dropdown,
    BoundedIntText,
    Label,
)
from openpyxl import load_workbook

from tabular_corpus import TabularCorpus, ParquetCorpus, serve_tabular_corpus


class UI:

    def __init__(self):

        button_layout = Layout(width="90%", height="2lh")
        wide_layout = Layout(width="90%")
        style = {"description_width": "40%"}

        self.upload_file = FileUpload(
            accept=".xlsx,.parquet",
            description="Upload your data (.xlsx, .parquet)",
            layout=button_layout,
        )

        # Incremental display components for progressing through a workflow
        self.table_config_output = Output()
        self.display_config_output = Output()
        self.process_output = Output()

        # Table config components
        self.table_label = Label(
            value="Table and column configuration", layout=wide_layout
        )
        self.table_select = Dropdown(
            description="Select table/sheet to use",
            layout=wide_layout,
            style=style,
        )
        self.text_column_select = SelectMultiple(
            description="Select the text columns to make searchable",
            layout=wide_layout,
            style=style,
        )

        # Display config components
        self.display_label = Label(
            value="Document display configuration", layout=wide_layout
        )
        self.display_style = Dropdown(
            options=["transcript", "document", "concordance"],
            description="Document display style",
            layout=wide_layout,
            style=style,
        )
        self.header_columns = SelectMultiple(
            description="Header columns for search result display",
            layout=wide_layout,
            style=style,
        )
        self.inline_columns = SelectMultiple(
            description="Inline columns with text for search result display",
            layout=wide_layout,
            style=style,
        )
        self.filter_columns = SelectMultiple(
            description="Columns to use for filtering results",
            layout=wide_layout,
            style=style,
        )
        self.context_turn_count = BoundedIntText(
            value=1,
            min=0,
            step=1,
            description="Number of context turns to show (transcripts only)",
            layout=wide_layout,
            style=style,
        )

        # Current uploaded file state
        self.current_file_type = None
        self.current_file = None
        self.table_cols = {}
        self.showing_config = False

        # Register callbacks for moving through the process
        self.upload_file.observe(self.create_table_config, names=["value"])
        self.table_select.observe(self.update_columns, names=["value"])
        self.text_column_select.observe(self.display_config, names=["value"])

        # The final run button
        self.run_button = Button(
            description="Explore your data table",
            layout=Layout(width="90%", height="2lh"),
        )
        self.run_button.on_click(self.run)

        self.table_config_components = VBox(
            [self.table_label, self.table_select, self.text_column_select]
        )

        self.display_config_components = VBox(
            [
                self.display_label,
                self.display_style,
                self.header_columns,
                self.inline_columns,
                self.filter_columns,
                self.context_turn_count,
                self.run_button,
            ]
        )

        # Container for the final output
        self.display_ui = VBox(
            [
                self.upload_file,
                self.table_config_output,
                self.display_config_output,
                self.process_output,
            ]
        )

        display(self.display_ui)

        self.corpus = None

    def run(self, clicked_button):
        # Cleanup corpus and any stray index files.
        for path in ("tabular_corpus.db", "corpus_index.db"):
            for extension in ("", "-shm", "-wal"):
                pathlib.Path(path + extension).unlink(missing_ok=True)

        with self.process_output:
            print(f"Preparing corpus for {self.upload_file.value[0].name}")

            if self.current_file_type == ".xlsx":
                self.create_corpus_from_excel()

            elif self.current_file_type == ".parquet":
                self.create_corpus_from_parquet()

            self.run_button.disabled = True
            serve_tabular_corpus(self.corpus)
            print("(Run the note book again `▶▶` if you want to restart)")

    def clear(self):
        """Clear all output components and working components."""
        for component in (
            self.table_config_output,
            self.display_config_output,
            self.process_output,
        ):
            component.clear_output()

        self.current_file_type = None
        self.current_file = None
        self.table_cols = {}

    def update_columns(self, change):
        self.text_column_select.options = self.table_cols.get(change.new, [])

    def create_corpus_from_excel(self):
        """Create the actual tabular corpus from the full set of documents."""

        if self.corpus is not None:
            self.corpus.close()

        text_columns = list(self.text_column_select.value)
        filter_columns = list(self.filter_columns.value)

        self.corpus = TabularCorpus(
            "tabular_corpus.db",
            text_fields=text_columns,
            filter_fields=filter_columns,
            header_fields=list(self.header_columns.value or []),
            inline_fields=list(self.inline_columns.value or []),
            display_style=str(self.display_style.value),
            display_transcript_context_turns=int(self.context_turn_count.value),
        )

        self.corpus.replace_rows_from_spreadsheet(
            self.current_file, self.table_select.value
        )

    def create_corpus_from_parquet(self):
        """Create the actual tabular corpus from the full set of documents."""

        if self.corpus is not None:
            self.corpus.close()

        text_columns = list(self.text_column_select.value)
        filter_columns = list(self.filter_columns.value)

        self.corpus = ParquetCorpus(
            "uploaded.parquet",
            text_fields=text_columns,
            filter_fields=filter_columns,
            header_fields=list(self.header_columns.value or []),
            inline_fields=list(self.inline_columns.value or []),
            display_style=str(self.display_style.value),
            display_transcript_context_turns=int(self.context_turn_count.value),
        )

    def create_excel_config(self):
        """Create the config just for an excel file."""

        wb = self.current_file
        sheetnames = wb.sheetnames

        # Sheetname, column mappings
        self.table_cols = {}

        for sheet in sheetnames:
            ws = wb[sheet]
            cols = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            self.table_cols[sheet] = [col or "" for col in cols]

        self.table_select.options = sheetnames

        if len(sheetnames) == 1:
            self.table_select.value = sheetnames[0]

    def create_parquet_config(self):
        """Create the config just for a parquet file."""

        # Make sure to skip the generated __docid column
        columns = list(polars.scan_parquet(self.current_file).collect_schema())[1:]
        self.table_cols = {"": columns}

        self.table_select.options = [""]
        self.table_select.value = ""
        self.table_select.disable = True

    def create_table_config(self, change):
        """Create the table configuration workflow from the provided file."""
        self.clear()

        with self.table_config_output:
            uploaded = change.new[0]

            uploaded_name = uploaded.name.lower()

            if uploaded_name.endswith(".xlsx"):
                self.current_file_type = ".xlsx"
                self.current_file = load_workbook(filename=BytesIO(uploaded.content))
                self.create_excel_config()

            elif uploaded_name.endswith(".parquet"):
                self.current_file_type = ".parquet"
                # Write out the uploaded parquet file with an additional named and
                # sorted index column: this is needed for fast lookups by row.
                df = polars.scan_parquet(BytesIO(uploaded.content)).with_row_index(
                    "__doc_id"
                )
                df.sink_parquet("uploaded.parquet", compression_level=22)
                self.current_file = "uploaded.parquet"
                self.create_parquet_config()

            self.table_select.disable = False

            display(self.table_config_components)

    def display_config(self, change):
        # Set display column selections based on the selected text columns
        table_columns = self.table_cols[self.table_select.value]
        text_columns = self.text_column_select.value
        not_selected_columns = [col for col in table_columns if col not in text_columns]

        self.header_columns.options = not_selected_columns
        self.inline_columns.options = not_selected_columns
        self.filter_columns.options = not_selected_columns
        self.filter_columns.value = not_selected_columns

        # Only display config settings if it's a valid config now and wasn't previously
        # showing.
        if change.new and not self.showing_config:
            with self.display_config_output:
                display(self.display_config_components)

            self.showing_config = True

        elif not change.new and self.showing_config:
            # If there's now no selected columns, don't show the config yet.
            self.display_config_output.clear_output()
            self.showing_config = False


if __name__ == "__main__":
    # Test this by manually setting the results of each action and ensuring something
    # happens at the end...
    pass
